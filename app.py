"""
App Streamlit - Dashboard SIGE Enturmação
Utiliza enturmacao.py para baixar dados e exibe em dashboard.
"""
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from io import BytesIO

import pandas as pd
import streamlit as st

# Caminho do diretório do projeto (onde está enturmacao.py e o CSV)
DIR_PROJETO = Path(__file__).resolve().parent
ARQUIVO_CSV = DIR_PROJETO / "Relatorio_SIGE_Corrigido.csv"

st.set_page_config(
    page_title="SIGE Enturmação",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Estilo para métricas e cards
st.markdown("""
<style>
    .metric-card {
        background: linear-gradient(135deg, #1e3a5f 0%, #2d5a87 100%);
        color: white;
        padding: 1rem 1.25rem;
        border-radius: 10px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        margin-bottom: 0.5rem;
    }
    .metric-card h3 { margin: 0; font-size: 0.85rem; opacity: 0.9; }
    .metric-card .valor { font-size: 1.75rem; font-weight: 700; }
    div[data-testid="stMetricValue"] { font-size: 1.5rem !important; }
</style>
""", unsafe_allow_html=True)


# Nomes das colunas iguais ao enturmacao.py (COLUNAS_FINAIS)
COLUNAS_NUMERICAS = ["Mat. Total", "Não Enturmados", "Enturmados", "Quantidade de Turmas", "Mat. Presencial", "Mat. Semipresencial"]


@st.cache_data(ttl=300)
def carregar_csv(caminho):
    """Carrega o CSV do relatório (cache 5 min). Usa os mesmos nomes de colunas do enturmacao.py."""
    if not caminho.exists():
        return None
    try:
        df = pd.read_csv(caminho, sep=";", encoding="utf-8-sig")
        for col in COLUNAS_NUMERICAS:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col].astype(str).str.replace(".", "", regex=False).str.replace("-", "0", regex=False), errors="coerce").fillna(0).astype(int)
        return df
    except Exception:
        return None


def extrair_etapa(desc):
    """Extrai a etapa (Creche, 1º Ano, etc.) da coluna Descricao."""
    if pd.isna(desc):
        return ""
    s = str(desc).strip()
    if "|" in s:
        parte = s.split("|")[0].strip()
        if " - " in parte:
            return parte.rsplit(" - ", 1)[-1].strip()
        return parte
    if s == "TOTAL SECRETARIA":
        return "TOTAL SECRETARIA"
    return "Total escola"


def gerar_xlsx(df, filtros_aplicados):
    """Gera um XLSX com os dados filtrados."""
    buffer = BytesIO()
    
    # Lê a data da extração do arquivo log
    ARQUIVO_LOG = DIR_PROJETO / "ultima_extracao.txt"
    data_extracao = ""
    if ARQUIVO_LOG.exists():
        try:
            data_extracao = ARQUIVO_LOG.read_text(encoding='utf-8').strip()
        except Exception:
            pass
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Relatório', index=False)
        
        # Ajusta a largura das colunas
        worksheet = writer.sheets['Relatório']
        for idx, col in enumerate(df.columns, 1):
            max_length = max(df[col].astype(str).map(len).max(), len(col))
            worksheet.column_dimensions[chr(64 + idx)].width = min(max_length + 2, 40)
        
        # Formata o cabeçalho
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col_num, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Coloriza as linhas conforme o Status
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        
        if 'Status' in df.columns:
            status_col_idx = list(df.columns).index('Status') + 1
            for row_num, status in enumerate(df['Status'], start=2):
                if status == 'Atenção':
                    fill = yellow_fill
                elif status == 'Crítica':
                    fill = red_fill
                else:
                    continue
                
                for col_num in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_num, column=col_num).fill = fill
        
        # Adiciona uma aba com informações dos filtros e data da extração
        ws_filtros = writer.book.create_sheet('Informações', 0)
        ws_filtros['A1'] = 'Relatório SIGE - Mapa de Enturmação'
        ws_filtros['A3'] = 'Data de Extração dos Dados:'
        ws_filtros['B3'] = data_extracao if data_extracao else "Não informada"
        ws_filtros['A5'] = 'Data do Relatório:'
        ws_filtros['B5'] = f'{datetime.now().strftime("%d/%m/%Y às %H:%M")}'
        
        if filtros_aplicados:
            ws_filtros['A7'] = 'Filtros Aplicados:'
            
            for idx, filtro in enumerate(filtros_aplicados, start=8):
                ws_filtros[f'A{idx}'] = f'• {filtro}'
    
    buffer.seek(0)
    return buffer


def gerar_xlsx_por_municipio(df, filtros_aplicados):
    """Gera um arquivo ZIP com XLSXs separados por município."""
    import zipfile
    
    # Lê a data da extração do arquivo log
    ARQUIVO_LOG = DIR_PROJETO / "ultima_extracao.txt"
    data_extracao = ""
    if ARQUIVO_LOG.exists():
        try:
            data_extracao = ARQUIVO_LOG.read_text(encoding='utf-8').strip()
        except Exception:
            pass
    
    zip_buffer = BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        municipios = df['Municipio'].dropna().unique()
        
        for municipio in sorted(municipios):
            df_mun = df[df['Municipio'] == municipio]
            
            # Cria Excel para o município
            xlsx_buffer = BytesIO()
            
            with pd.ExcelWriter(xlsx_buffer, engine='openpyxl') as writer:
                df_mun.to_excel(writer, sheet_name='Relatório', index=False)
                
                # Ajusta a largura das colunas
                worksheet = writer.sheets['Relatório']
                for idx, col in enumerate(df_mun.columns, 1):
                    max_length = max(df_mun[col].astype(str).map(len).max(), len(col))
                    worksheet.column_dimensions[chr(64 + idx)].width = min(max_length + 2, 40)
                
                # Formata o cabeçalho
                from openpyxl.styles import PatternFill, Font, Alignment
                header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')
                
                for col_num, col_name in enumerate(df_mun.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Coloriza as linhas conforme o Status
                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                
                if 'Status' in df_mun.columns:
                    for row_num, status in enumerate(df_mun['Status'], start=2):
                        if status == 'Atenção':
                            fill = yellow_fill
                        elif status == 'Crítica':
                            fill = red_fill
                        else:
                            continue
                        
                        for col_num in range(1, len(df_mun.columns) + 1):
                            worksheet.cell(row=row_num, column=col_num).fill = fill
                
                # Adiciona aba com informações
                ws_info = writer.book.create_sheet('Informações', 0)
                ws_info['A1'] = f'Relatório SIGE - {municipio}'
                ws_info['A3'] = 'Data de Extração dos Dados:'
                ws_info['B3'] = data_extracao if data_extracao else "Não informada"
                ws_info['A5'] = 'Data do Relatório:'
                ws_info['B5'] = f'{datetime.now().strftime("%d/%m/%Y às %H:%M")}'
                
                if filtros_aplicados:
                    ws_info['A7'] = 'Filtros Aplicados:'
                    for idx, filtro in enumerate(filtros_aplicados, start=8):
                        ws_info[f'A{idx}'] = f'• {filtro}'
            
            xlsx_buffer.seek(0)
            nome_arquivo = f"{municipio}.xlsx"
            zip_file.writestr(nome_arquivo, xlsx_buffer.getvalue())
    
    zip_buffer.seek(0)
    return zip_buffer


def executar_enturmacao():
    """Roda enturmacao.py e retorna (sucesso, mensagem)."""
    try:
        proc = subprocess.run(
            [sys.executable, str(DIR_PROJETO / "enturmacao.py")],
            cwd=str(DIR_PROJETO),
            capture_output=True,
            text=True,
            timeout=600,
            encoding="utf-8",
            errors="replace",
        )
        if proc.returncode == 0:
            return True, "Dados baixados com sucesso."
        return False, proc.stderr or proc.stdout or "Erro ao executar script."
    except subprocess.TimeoutExpired:
        return False, "Tempo esgotado (o download pode demorar vários minutos)."
    except Exception as e:
        return False, str(e)


def main():
    st.title("📊 Relatório SIGE - Mapa de Enturmação")
    st.caption("Dados do relatório de enturmação por secretaria, município e escola.")

    # Sidebar: download e filtros
    with st.sidebar:
        st.header("Dados")
        if st.button("🔄 Baixar dados do SIGE", type="primary", use_container_width=True):
            with st.spinner("Executando enturmacao.py... (pode demorar alguns minutos)"):
                ok, msg = executar_enturmacao()
                if ok:
                    st.success(msg)
                    st.cache_data.clear()
                else:
                    st.error(msg)

        st.divider()
        st.header("Filtros do dashboard")

    df = carregar_csv(ARQUIVO_CSV)

    if df is None or df.empty:
        st.warning("Nenhum dado carregado. Clique em **Baixar dados do SIGE** na barra lateral ou coloque o arquivo `Relatorio_SIGE_Corrigido.csv` na pasta do projeto.")
        st.info("O dashboard será exibido quando houver o arquivo Relatorio_SIGE_Corrigido.csv gerado pelo enturmacao.py (colunas: Secretaria, Municipio, Descricao, Mat. Total, Não Enturmados, Enturmados, Quantidade de Turmas, Mat. Presencial, Mat. Semipresencial, Status).")
        return

    # Data/hora da última atualização do CSV (lê do arquivo ultima_extracao.txt)
    ARQUIVO_LOG = DIR_PROJETO / "ultima_extracao.txt"
    if ARQUIVO_LOG.exists():
        try:
            data_atualizacao = ARQUIVO_LOG.read_text(encoding='utf-8').strip()
            st.caption(f"📅 Dados atualizados em: **{data_atualizacao}**")
        except Exception:
            if ARQUIVO_CSV.exists():
                mtime = ARQUIVO_CSV.stat().st_mtime
                data_atualizacao = datetime.fromtimestamp(mtime).strftime("%d/%m/%Y %H:%M")
                st.caption(f"📅 Dados atualizados em: **{data_atualizacao}**")
    elif ARQUIVO_CSV.exists():
        mtime = ARQUIVO_CSV.stat().st_mtime
        data_atualizacao = datetime.fromtimestamp(mtime).strftime("%d/%m/%Y %H:%M")
        st.caption(f"📅 Dados atualizados em: **{data_atualizacao}**")

    # Coluna Etapa extraída da Descricao (para filtro)
    df = df.copy()
    df["Etapa"] = df["Descricao"].map(extrair_etapa)

    # Filtros na sidebar — sem Secretaria; Etapa e Status são multiselect
    with st.sidebar:
        municipios = ["Todos os Municípios"] + sorted(df["Municipio"].dropna().unique().tolist())
        sel_municipio = st.selectbox("Município", municipios)
        opcoes_etapa = sorted(e for e in df["Etapa"].dropna().unique() if str(e).strip())
        sel_etapas = st.multiselect("Etapa", opcoes_etapa, default=[], placeholder="Todas")
        status_opcoes = sorted(df["Status"].dropna().unique().tolist())
        sel_status = st.multiselect("Status", status_opcoes, default=[], placeholder="Todos")

    # Aplicar filtros
    mask = pd.Series(True, index=df.index)
    if sel_municipio != "Todos os Municípios":
        mask &= df["Municipio"] == sel_municipio
    if sel_etapas:
        mask &= df["Etapa"].isin(sel_etapas)
    if sel_status:
        mask &= df["Status"].isin(sel_status)
    df_filtrado = df[mask]

    st.divider()
    
    # Exibir hora da extração em destaque
    ARQUIVO_LOG = DIR_PROJETO / "ultima_extracao.txt"
    if ARQUIVO_LOG.exists():
        try:
            data_extracao = ARQUIVO_LOG.read_text(encoding='utf-8').strip()
            col_info1, col_info2, col_info3 = st.columns([1, 2, 1])
            with col_info2:
                # st.info(f"⏰ **Última extração:** {data_extracao}", icon="ℹ️")
                pass
        except Exception:
            pass

    # Tabela completa (sem coluna Secretaria; nomes das colunas = enturmacao.py)
    st.subheader("Dados completos (filtrados)")
    df_exibir = df_filtrado.drop(columns=["Secretaria"], errors="ignore")
    
    # Função para colorizar as linhas conforme o Status
    def pintar_status(row):
        if 'Status' in df_exibir.columns:
            status = row['Status']
            if status == "Atenção":
                return ['background-color: #FFFF00'] * len(row)  # Amarelo
            elif status == "Crítica":
                return ['background-color: #FF6B6B'] * len(row)  # Vermelho
        return [''] * len(row)
    
    df_styled = df_exibir.style.apply(pintar_status, axis=1)
    st.dataframe(df_styled, use_container_width=True, height=400)

    # Botão para download de PDF e Excel
    st.divider()
    
    # Prepara informações dos filtros
    filtros_info = []
    if sel_municipio != "Todos os Municípios":
        filtros_info.append(f"Município: {sel_municipio}")
    if sel_etapas:
        filtros_info.append(f"Etapa(s): {', '.join(sel_etapas)}")
    if sel_status:
        filtros_info.append(f"Status: {', '.join(sel_status)}")
    
    # Layout diferente conforme o município selecionado
    if sel_municipio == "Todos os Municípios":
        # Mostrar botão para Excel Consolidado e por Município
        st.subheader("Downloads")
        
        col1, col2 = st.columns([1, 1])
        
        with col1:
            xlsx_buffer = gerar_xlsx(df_exibir, filtros_info)
            st.download_button(
                label="📊 Excel Consolidado",
                data=xlsx_buffer,
                file_name=f"Relatorio_SIGE_Todos_{datetime.now().strftime('%d%m%Y_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        
        with col2:
            zip_xlsx_buffer = gerar_xlsx_por_municipio(df_exibir, filtros_info)
            st.download_button(
                label="📋 Excel (por Município)",
                data=zip_xlsx_buffer,
                file_name=f"Relatorio_SIGE_Municipios_Excel_{datetime.now().strftime('%d%m%Y_%H%M%S')}.zip",
                mime="application/zip",
                use_container_width=True
            )
    else:
        # Mostrar botão de Excel quando um município específico está selecionado
        col1, col2, col3 = st.columns([1, 1, 1])
        
        with col2:
            xlsx_buffer = gerar_xlsx(df_exibir, filtros_info)
            st.download_button(
                label="📊 Excel",
                data=xlsx_buffer,
                file_name=f"Relatorio_SIGE_{sel_municipio}_{datetime.now().strftime('%d%m%Y_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )


if __name__ == "__main__":
    main()
